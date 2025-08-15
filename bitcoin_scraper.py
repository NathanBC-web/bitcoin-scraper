from pycoingecko import CoinGeckoAPI
import pandas as pd
from datetime import datetime
import os
import logging
from fredapi import Fred
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    filename='bitcoin_scraper.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

try:
    logging.info("Starting Bitcoin price scraping process...")
    cg = CoinGeckoAPI()

    # Initialize FRED API with your API key
    fred = Fred(api_key=os.getenv('FRED_API_KEY'))  # Using environment variable

    # Get Bitcoin price data for the last 120 days
    logging.info("Fetching Bitcoin price data from CoinGecko...")
    data = cg.get_coin_market_chart_by_id(id='bitcoin', vs_currency='usd', days=120)

    # Convert timestamps to readable dates (without time) and collect prices
    logging.info("Processing price data into DataFrame...")
    prices = [(datetime.fromtimestamp(p[0]/1000).strftime('%Y-%m-%d'), p[1], None, None) for p in data['prices']]
    df = pd.DataFrame(prices, columns=['Date', 'Bitcoin Price', 'Global Liquidity (M2)', 'Expected Bitcoin Price'])

    # Get M2 data from FRED (last 365 days for better data availability)
    logging.info("Fetching M2 data from FRED...")
    end_date = datetime.now().strftime('%Y-%m-%d')
    start_date = (datetime.now() - pd.Timedelta(days=365)).strftime('%Y-%m-%d')  # Extended to 1 year
    m2_data = fred.get_series('M2SL', observation_start=start_date, observation_end=end_date)

    # Convert M2 index to datetime and interpolate to daily (average approximation)
    m2_df = m2_data.reset_index()
    m2_df.columns = ['Date', 'Global Liquidity (M2)']
    m2_df['Date'] = pd.to_datetime(m2_df['Date'])
    m2_df = m2_df.set_index('Date').resample('D').interpolate(method='linear').reset_index()
    m2_df['Date'] = m2_df['Date'].dt.strftime('%Y-%m-%d')  # Format as string without time

    # Group by date and calculate the average price per day, preserving other columns
    logging.info("Calculating daily average prices...")
    daily_avg_df = df.groupby('Date', as_index=False).agg({'Bitcoin Price': 'mean', 'Global Liquidity (M2)': 'first', 'Expected Bitcoin Price': 'first'})
    daily_avg_df['Date'] = pd.to_datetime(daily_avg_df['Date'])  # Convert to datetime
    daily_avg_df['Date'] = daily_avg_df['Date'].dt.strftime('%Y-%m-%d')  # Format as string without time

    # Drop the original empty Global Liquidity (M2) column before merging
    daily_avg_df = daily_avg_df.drop(columns=['Global Liquidity (M2)'])

    # Merge M2 data with Bitcoin data
    logging.info("Merging M2 data with Bitcoin data...")
    daily_avg_df = daily_avg_df.merge(m2_df, on='Date', how='left')

    # Reorder columns to have Expected Bitcoin Price as the last column
    logging.info("Reordering columns...")
    daily_avg_df = daily_avg_df[['Date', 'Bitcoin Price', 'Global Liquidity (M2)', 'Expected Bitcoin Price']]

    # Sort by date descending (newest first)
    logging.info("Sorting data by date (newest first)...")
    daily_avg_df = daily_avg_df.sort_values(by='Date', ascending=False)

    # Define output path with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = f'C:\\Users\\ArhanPeek\\OneDrive - Universal IT B.V\\Documents\\Crypto\\bitcoin_prices_{timestamp}.xlsx'

    # Create folder if it doesn't exist
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Save to Excel with column width and frozen header
    logging.info(f"Saving data to Excel file: {output_path}")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        daily_avg_df.to_excel(writer, sheet_name='Bitcoin Prices', index=False)
        
        # Adjust column widths
        workbook = writer.book
        worksheet = writer.sheets['Bitcoin Prices']
        for idx, column in enumerate(['Date', 'Bitcoin Price', 'Global Liquidity (M2)', 'Expected Bitcoin Price']):
            # Set base length from column name
            base_length = len(str(column))
            # Fixed width for Date column (12 for "YYYY-MM-DD" with padding)
            if column == 'Date':
                max_length = 12
            # For other columns, use max of name or data length; minimum 25 for empty columns
            elif daily_avg_df[column].isna().all():
                max_length = max(base_length, 25)  # Minimum 25 for long empty titles
            else:
                max_length = max(base_length, daily_avg_df[column].astype(str).str.len().max())
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2

        # Freeze the header row (row 1)
        worksheet.freeze_panes = 'A2'

        # Apply formatting to Global Liquidity (M2) column
        m2_original_dates = m2_data.index.strftime('%Y-%m-%d').tolist()  # Original FRED dates
        for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):  # Column 3 is Global Liquidity (M2)
            for cell in row:
                if pd.notna(cell.value) and str(cell.offset(column=-2).value) in m2_original_dates:
                    cell.font = Font(bold=True)  # Bold for original data
                elif pd.notna(cell.value):
                    cell.font = Font(color="808080")  # Gray text for interpolated data

        # Freeze the header row (row 1)
        worksheet.freeze_panes = 'A2'

    print(f"✅ Daily average price Excel file saved successfully at:\n{output_path}")
    logging.info("Bitcoin price scraping process completed successfully.")

except Exception as e:
    error_message = f"An error occurred: {str(e)}"
    print("❌ " + error_message)
    logging.error(error_message)

# Pause so the window stays open
input("\nPress Enter to close...")